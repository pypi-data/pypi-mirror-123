"""
Title:
Author: Julio Rodriguez
Organization: SENER
"""

# BLK: Imports

# Standard Libraries
from datetime import datetime, timedelta
import sys
import time
import os
import io
from multiprocessing import Process, Pool, Semaphore, cpu_count
import subprocess
from pathlib import Path
from functools import wraps
import pickle
import cProfile
import pstats


# 3rd Party
# import pandas as pd
import numpy as np
import openpyxl.utils.exceptions
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook


# Local Libraries
from . import solar_common as _cm


# Global variables
class Cstruct:
    pass


# Code

# Clases

class Template:
    def __init__(self, template_file):
        self.original = template_file

        # Output data
        self.nombres_columna_general = None
        self.pos_loc = None
        self.sectores_full = None
        self.kks_loops_full = None
        self.maquinas_totales = None
        self.column = None
        self.lat = None
        self.lon = None
        self.row = None
        self.ew_sect = None
        self.ring = None
        self.tot_locs = None
        self.kks_loops = None
        self.index_kks_loops = None
        self.sectores = None
        self.sect_div = None
        self.n_lazos_sect = None

    def get_data(self):
        self.nombres_columna_general = np.array(self.original.columns, dtype='U140')  # Cargo las columnas de la plantilla de la planta para poder saber cuando empiezan los datos
        self._column_data()
        self._general_data()
        self._uniques_lazos()
        self._sector_calc()

    def _column_data(self):
        # Column data
        self.pos_loc = np.array(self.original["LOC POSITION IN THE LOOP"], dtype=int)
        self.column = np.array(self.original["COLUMN"], dtype=int)
        self.lat = np.array(self.original["Latitude (º)"], dtype=float)
        self.lon = np.array(self.original["Longitude (º)"], dtype=float)
        self.row = np.array(self.original["ROW"])
        self.ew_sect = np.array(self.original["E/W"])
        self.ring = np.array(self.original["RING / SUBFIELD"], dtype=int)

    def _general_data(self):
        # General data
        self.tot_locs = self.original.shape[0]
        self.sectores_full = np.array(self.original["HTF SECTOR"], dtype='U140')
        self.kks_loops_full = np.array(self.original["KKS LOOP"], dtype='U140')  # Array con los loops [numero de locs]
        self.maquinas_totales = np.array(self.original["KKS/Variable"], dtype='U140')  # Array con los nombres de los LOCs [numero de locs]

    def _uniques_lazos(self):
        # sBLK: Uniques de los lazos, máquinas y sectores
        # Unique sirve para obtener los valores sin repetir del array pasado
        # return_inverse devuelve un array de shape como el original, pero con los valores que indican la posición del array sin repetir (para poder re-estructurar el array original)
        self.kks_loops, self.index_kks_loops = np.unique(self.kks_loops_full, return_inverse=True)  # kks_loops [numero de lazos], index_kks_loops [numero de locs]
        self.sectores = pd.unique(self.sectores_full)  # Unique de los sectores para saber que sectores hay [numero de sectores]

    def _sector_calc(self):
        # sBLK: Dividir kks loops por sectores
        # Obtener el sector al que pertenece cada lazo (kks_loops)
        sectores_loop = np.empty(0)
        for i in np.arange(self.kks_loops.size):
            pos_sector = np.where(self.kks_loops_full == self.kks_loops[i])[0][0]  # Donde los kks_loops de el archivo sean el kks de i. Se accede con [0] al tuple y luego otro [0] para el primer elemento del array
            sectores_loop = np.append(sectores_loop, self.sectores_full[pos_sector])  # Se guarda el sector que hay en la posicion de uno de los LOCs

        # Crear un array que contiene las posiciones en kks_loops en donde se cambia de sector (ya que la plantilla está ordenada por sector)
        # FIXME: Puede que en alguna planta no esté siempre ordenado por sector, buscar una solución a eso (ordenar por sectores nada más cargar por ejemplo)
        self.sect_div = _cm.array_division(sectores_loop) + 1  # Posiciones de la division de los loops por sectores
        self.sect_div = np.insert(self.sect_div, 0, 0)  # Inserto un 0 al principio como posición incial

        # Calcular el nº de lazos por sector
        self.n_lazos_sect = np.zeros(self.sect_div.size - 1, dtype=int)
        for k in np.arange(self.sect_div.size - 1):
            self.n_lazos_sect[k] = self.sect_div[k + 1] - self.sect_div[k]


# Funciones


def crear_directorio_v2(ruta_defecto, nombre):
    """
    A diferencia de la versión 1, esta función no necesita que se le pase el nombre con la barra '/' delante, realiza un join internamente.

    :param ruta_defecto: string
    Ruta base donde se va a crear el directorio

    :param nombre: string
    Nombre del directorio nuevo

    :return: string
    Devuelve la ruta creada

    """
    ruta_final = os.path.join(ruta_defecto, nombre)
    if not os.path.exists(ruta_final):
        os.mkdir(ruta_final)
    return ruta_final


def porcentaje_x(array, tam, axis=0):
    array = np.true_divide(np.sum(array, axis=axis), tam) * 100
    return array


def consecutive(data, stepsize=1):
    return np.split(data, np.where(np.diff(data) != stepsize)[0] + 1)


def convert_to_float(frac_str):
    '''
    Convierte un string que contiene una fracción en numero decimal con el que operar en Python.
    :param frac_str: String que contiene la fracción.
    :return: Fracción convertida en numero de tipo float.
    '''
    try:
        return float(frac_str)
    except ValueError:
        num, denom = frac_str.split('/')
        try:
            num_temp = num
            leading, num = num_temp.split(' ')
            whole = float(leading)
        except ValueError:
            try:
                float(num)
                whole = 0
            except ValueError:
                if num_temp[0] == '-':
                    temp, leading, num = num_temp.split(' ')
                    whole = -float(leading)
                    num = '-' + num

        frac = float(num) / float(denom)
        return whole - frac if whole < 0 else whole + frac


def estimar_datos(datos, num_datos, tolerancia):
    suma = np.true_divide(np.sum(datos, axis=0), num_datos)
    estimacion = np.where(suma > tolerancia, 1, 0)
    return estimacion


def appendear_datos(first_data, axis, *args):
    try:
        first_data.shape[axis]
    except IndexError:
        first_data = first_data[:, np.newaxis]
    finally:
        out = first_data
    for data in args:
        try:
            data.shape[axis]
        except IndexError:
            try:
                data = data[:, np.newaxis]
            except IndexError:
                data = [data]
        finally:
            out = np.append(out, data, axis=axis)
    return out


def obtener_x_maxmin(data, n_elementos=None, orden='max'):
    """
    Obtiene los n_elementos máximos (max) o mínimos (min) de un ndarray.

    :param data: Datos a ser ordenados (ndarray)
    :param n_elementos: Nº máx. de elementos a devolver
    :param orden: De menor a mayor(min) o de mayor a menor(max)
    :return: Array data ordenado y array con los indices de ordenado
    """
    if n_elementos is None:
        n_elementos = data.size
    # Obtener los peores lazos (los que tengan mayor valor o menor dependiendo de abrir (negativo) o cerrar (positivo))
    # ind_tmp = np.argpartition(data, -n_elementos)[-n_elementos:]
    if orden == 'min':
        # ind_tmp_ord = ind_tmp[np.argsort(data[ind_tmp])]
        ind_tmp_ord = np.argsort(data)[:n_elementos]
    elif orden == 'max':
        # ind_tmp_ord = np.flip(ind_tmp[np.argsort(data[ind_tmp])])
        ind_tmp_ord = np.argsort(data)[-n_elementos:]
    else:
        raise ValueError("El valor introducido no es aceptado. Utiliza 'min' para obtener los mínimos y 'max' para obtener los máximos")

    return data[ind_tmp_ord], ind_tmp_ord


def multiprocesamiento(function, n_procesos, max_procesos=cpu_count(), *args):
    sema = Semaphore(max_procesos)
    all_processes = []
    for i in range(n_procesos):
        # once 20 processes are running, the following `acquire` call
        # will block the main process since `sema` has been reduced
        # to 0. This loop will continue only after one or more
        # previously created processes complete.
        sema.acquire()
        p = Process(target=function, args=(args, sema))
        all_processes.append(p)
        p.start()

    # inside main process, wait for all processes to finish
    for p in all_processes:
        p.join()


def generar_video_mp4(fig, fps, update, nom_video, n_frames, *args, **kwargs):
    # Inicialización de kwargs
    kwargs["returns"] = kwargs.get("returns", None)

    if type(fps) != str:
        fps = str(fps)

    canvas_width, canvas_height = fig.canvas.get_width_height()

    # Open an ffmpeg proce
    outf = nom_video
    cmdstring = ('ffmpeg',
                 '-y', '-r', fps,  # overwrite, x fps
                 '-s', '%dx%d' % (canvas_width, canvas_height),  # size of image string
                 '-pix_fmt', 'argb',  # format
                 '-f', 'rawvideo', '-i', '-',  # tell ffmpeg to expect raw video from the pipe
                 '-hide_banner', '-loglevel', 'error',   # Hides the description of the options selected
                 '-vcodec', 'mpeg4',
                 '-b:v', '5000k', outf)  # output encoding
    p = subprocess.Popen(cmdstring, stdin=subprocess.PIPE, shell=True)

    # Draw 1000 frames and write to the pipe
    for frame in tqdm(np.arange(n_frames)):

        # draw the frame, PONER AQUI LA FUNCIÓN QUE DIBUJA
        kwargs["returns"] = update(fig, frame, *args, **kwargs)
        fig.canvas.draw()

        # extract the image as an ARGB string
        string = fig.canvas.tostring_argb()

        # write to pipe
        p.stdin.write(string)

        # Clear the figure
        fig.clf()

    # Finish up
    p.communicate()
    return kwargs["returns"]


def convertir_df_numerico(df, columna):
    df[columna] = df[columna].apply(pd.to_numeric)
    return df


def ajustar_columnas_excel(writer, nombre_hoja, n_col, data, index=None, sectores_extendido=False, resumen=False, normal=False, fila=0, col_off=0):
    # Ajusta el ancho de las columnas de un archivo excel
    worksheet = writer.sheets[nombre_hoja]
    # width = np.max([np.max([len(index.values)]), len(data.names[0])]) + 2
    if normal:
        for i in np.arange(n_col):
            width = len(data.values[i]) + 2
            worksheet.set_column(fila, i + col_off, width)
    else:
        width = len(data.names[0]) + 2  # Primera columna
        worksheet.set_column(0, 0, width)
        if sectores_extendido:
            width = np.max([len(data.values[0][0]), len(data.values[0][1])]) + 2  # Máximo entre la fecha y la alarma
            worksheet.set_column(1, 1, width)
            width = np.max([len(data.values[0][0]), len(data.values[1][1])]) + 2  # Máximo entre la fecha y la máquina
            worksheet.set_column(2, 2, width)
            for i in np.arange(2, n_col, 2):
                width1 = len(data.values[i][1]) + 2
                width2 = len(data.values[i + 1][1]) + 2
                worksheet.set_column(i + 1, i + 1, width1)
                worksheet.set_column(i + 2, i + 2, width2)
        elif resumen:
            for i in np.arange(n_col - 1) + 1:  # El +1 -1 es porque el primer elemento lo meto arriba como names, pero podría pasarse de el y poner sin -1 +1.
                try:
                    if data.values[i][0] != data.values[i + 1][0] and data.values[i][0] != data.values[i - 1][0]:  # Este if mira a ver si seguimos en la misma multicolumna, y uso el try porque en la última va a dar un error de out of index
                        width = len(data.values[i][0]) + 2
                    else:
                        width = len(data.values[i][1]) + 2
                except:
                    if i == 0:  # Este error no debería saltar ya que el for empieza con el primer elemento = 1
                        if data.values[i][0] != data.values[i + 1][0]:
                            width = len(data.values[i][0]) + 2
                        else:
                            width = len(data.values[i][1]) + 2
                    elif i == n_col - 1:  # Este error salta cuando llega a la última posicion el for ya que no existe última + 1, por lo que la realiza dentro de esta excepción
                        if data.values[i][0] != data.values[i - 1][0]:
                            width = len(data.values[i][0]) + 2
                        else:
                            width = len(data.values[i][1]) + 2
                finally:
                    worksheet.set_column(i, i, width)

        else:  # Cuando no tiene ningun multiindex, hace un ajuste normal
            for i in np.arange(n_col) + 1:
                width = len(data.values[i - 1][0]) + 4
                worksheet.set_column(i, i, width)


def string_to_time(a_tiempo_str):
    t_alarma_tot = np.full(a_tiempo_str.size, timedelta(days=0, hours=0, minutes=0, seconds=0), dtype=timedelta)
    a_tiempo_str = np.char.replace(a_tiempo_str, "/", "-")
    try:
        datetime.strptime(a_tiempo_str[0], '%H:%M:%S')
    except ValueError:
        try:
            datetime.strptime(a_tiempo_str[0], '%Y-%m-%d %H:%M:%S')
        except:
            for k in np.arange(a_tiempo_str.size):
                tiempo_tmp = datetime.strptime(a_tiempo_str[k], '%d-%m-%Y %H:%M:%S')
                t_alarma_tot[k] = t_alarma_tot[k] + timedelta(hours=tiempo_tmp.hour, minutes=tiempo_tmp.minute, seconds=tiempo_tmp.second)
        else:
            for k in np.arange(a_tiempo_str.size):
                tiempo_tmp = datetime.strptime(a_tiempo_str[k], '%Y-%m-%d %H:%M:%S')
                t_alarma_tot[k] = t_alarma_tot[k] + timedelta(hours=tiempo_tmp.hour, minutes=tiempo_tmp.minute, seconds=tiempo_tmp.second)
    else:
        for k in np.arange(a_tiempo_str.size):
            tiempo_tmp = datetime.strptime(a_tiempo_str[k], '%H:%M:%S')
            t_alarma_tot[k] = t_alarma_tot[k] + timedelta(hours=tiempo_tmp.hour, minutes=tiempo_tmp.minute, seconds=tiempo_tmp.second)
    return t_alarma_tot


def find_data_file(filename):
    if getattr(sys, 'frozen', False):
        # The application is frozen
        datadir = os.path.dirname(sys.executable)
    else:
        # The application is not frozen
        # Change this bit to match where you store your data files:
        datadir = os.path.dirname(__file__)
    return os.path.join(datadir, filename)


def read_corrupt(filename, sheet_name='Hoja1', skiprows=0, df_return=False):

    # Nombre fichero de salida
    # Se realiza un split para añadir _FIX al final del nombre del archivo
    f_split = filename.split('.')
    # Crear subcarpeta para meter ahí los ficheros modificados
    crear_directorio_v2(str(Path(f_split[0]).parent), 'Fixed')
    # Crear el nombre del fichero en la ruta 'Fixed' con el mismo nombre que el original pero extensión .xlsx
    filename_out = str(Path(f_split[0]).parent / 'Fixed' / f'{Path(f_split[0]).name}.xlsx')

    # Abrir fichero utilizando codificación 'utf-16'
    fich = io.open(filename, "r", encoding="utf-16")
    data = fich.readlines()

    # Abrir el fichero de salida y crear una página "Data"
    workbook = Workbook()
    sheet = workbook.create_sheet(sheet_name)

    # Iterar sobre los datos y guardarlos en Excel
    for row, fila in enumerate(data):
        # Quitar el '\n' que se inserta cuando se lee con el reader de io.open.
        # Cojer los valores despues de splittear con '\t'
        for column, val in enumerate(fila.replace('\n', '').split('\t')):
            try:
                sheet.cell(row + 1, column + 1, val)
            except openpyxl.utils.exceptions.IllegalCharacterError:
                print(f"Invalid line {row + 1}")
                break
    # Guardar el Excel (se genera)
    workbook.save(filename_out)

    if df_return:
        # Leer el fichero bueno con Pandas
        df_out = pd.read_excel(filename_out, sheet_name=sheet_name, skiprows=skiprows, engine='openpyxl')
    else:
        df_out = None

    return df_out


def generar_excel_default(data, nomfich, hoja="Hoja1", data_col=None, df_general=False, plant=None, nombres_col_gen=None, df_num=None, startrow=0, startcol=0, autofilter=False, index=False, header=True):
    writer = pd.ExcelWriter(nomfich, engine='openpyxl')
    if isinstance(data, np.ndarray):
        if data_col is not None:
            data = pd.DataFrame(data, columns=data_col)
        else:
            raise Exception("With a numpy ndarray is necessary to include data_col")

    if df_num is not None:
        for df in df_num:
            convertir_df_numerico(data, df)

    if plant is not None and nombres_col_gen is not None:
        _cm.convertir_df_numerico_GENERAL(data, nombres_col_gen, plant)

    data.to_excel(writer, sheet_name=hoja, startrow=startrow, startcol=startcol, index=False, header=True)

    # Activar autofilter
    if autofilter:
        worksheet = writer.sheets[hoja]
        worksheet.auto_filter.ref = worksheet.dimensions
        # worksheet.autofilter(0, 0, data.shape[0], data_col.size)  # Se usaría en caso de utilizar xlsxwriter

    writer.save()


def time_calculator(func, *args, **kwargs):
    start = time.time_ns()
    func(*args, **kwargs)
    print(rf"Time: {time.time_ns() - start}")


def find_adjacent_loops(Data, kks_loops, loop, n_adj, front_subfield=True):
    """
    Finds the adjacent loops of a loop in a solar field.

    :param kks_loops: Tags of all the loops at the Solar Field
    :param loop: Target loop
    :param n_adj: Number of adjacent loops to look for. Minimun value is 1 (front loop only would be)

    :param front_subfield: Look for adjacent loops in the subfield that shares with the current subfield. True by default
    :return: Array of 'n_adj' adjacent loops to target loop 'loop'
    """
    if n_adj < 1:
        raise Exception("The number of adjacent loops must be at least 1")

    pos_target = np.argwhere(np.array(np.logical_and(Data.kks_loops_full == loop, Data.pos_loc == 1)))[0, 0]

    # target = plantilla.iloc[pos_target]

    # Numero de adyacentes
    col_target = []
    for k in range(n_adj):
        col_target.append(Data.column[pos_target] + k * 2)
        col_target.append(Data.column[pos_target] - k * 2)

    # Cáclulo de adyacentes
    adj_loops = np.argwhere((Data.sectores_full == Data.sectores_full[pos_target]) &
                            (np.isin(Data.column, col_target)) &
                            (Data.pos_loc == 1)).squeeze(1)

    adj_loops = [np.argwhere(kks_loops == Data.kks_loops_full[loop])[0, 0] for loop in adj_loops]

    return adj_loops


def tag_splitter(tag):
    TAG = Cstruct()
    if hasattr(tag, "__len__"): # Comprueba si tiene el atributo __len__,
        # Array
        pass
    else:
        # String
        TAG.ew_sector = tag[3]


def delta_calculation(data, axis=0, nlocs=4):
    if data.ndim > 1:
        out_data = np.empty([data.shape[0], data.shape[1] - 1], dtype=float)
        for k in range(data.shape[axis]):
            for i in range(nlocs - 1):
                out_data[k, i] = data[k, i] - data[k, i + 1]
    else:
        out_data = np.empty(data.size - 1, dtype=float)
        for i in range(nlocs - 1):
            out_data[i] = data[i] - data[i + 1]

    return out_data


def negate(f):
    @wraps(f)
    def g(*args,**kwargs):
        return not f(*args,**kwargs)
    g.__name__ = f'negate({f.__name__})'
    return g


def guardar_var(ruta, var):  # Guarda una variable en la ruta indicada
    f = open(ruta, 'wb')
    pickle.dump(var, f)
    f.close()


def cargar_var(ruta):  # Devuelve el valor de una variable cargada desde la ruta indicada
    f = open(ruta, 'rb')
    var = pickle.load(f)
    f.close()
    return var
