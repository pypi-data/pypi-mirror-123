"""
Refer to the answer:
https://stackoverflow.com/questions/59379915/how-can-i-append-a-csv-file-to-an-existing-zip-archive-in-python
"""

from openpyxl import Workbook, load_workbook
import zipfile
import json
import csv
import io


zip_path = "/Users/bryantliu/Downloads/example_data_group.zip"
csv_path = "example_data.csv"
# xlsx_path = "/Users/bryantliu/Downloads/test_zip/test.xlsx"
xlsx_path = "example_data.xlsx"
json_path = "example_data.json"
csv_data = [
    ['0108-01-02 00:00:00', 32900482.0, 7276419230.0, 226.5, 226.5, 219.0, 219.5, 12329.0],
    ['0108-01-03 00:00:00', 34615620.0, 7459051790.0, 214.0, 218.0, 214.0, 215.5, 14549.0],
    ['0108-01-04 00:00:00', 67043521.0, 13987136785.0, 211.5, 211.5, 206.5, 208.0, 28786.0]
]
json_data = {"data": [
    ['0108-01-02 00:00:00', 32900482.0, 7276419230.0, 226.5, 226.5, 219.0, 219.5, 12329.0],
    ['0108-01-03 00:00:00', 34615620.0, 7459051790.0, 214.0, 218.0, 214.0, 215.5, 14549.0],
    ['0108-01-04 00:00:00', 67043521.0, 13987136785.0, 211.5, 211.5, 206.5, 208.0, 28786.0]
]}

# json object
json_data_str = json.dumps(json_data, ensure_ascii=False)


# csv object
string_io = io.StringIO()
csv_writer = csv.writer(string_io)
for d in csv_data:
    print(fr"d: {d}")
    csv_writer.writerow(d)
string_io.seek(0)


# # excel object
# # workbook = Workbook()
# workbook = load_workbook(filename=xlsx_path)
# sheet_page = workbook.create_sheet(index=0, title="example_data")
# for d in csv_data:
#     sheet_page.append(d)
# # workbook.save(xlsx_path)


# # zip
test_zip = zipfile.ZipFile(file=zip_path, mode="a", compression=zipfile.ZIP_DEFLATED, allowZip64=False)
test_zip.writestr(zinfo_or_arcname=json_path, data=json_data_str)
test_zip.writestr(zinfo_or_arcname=csv_path, data=string_io.read())
# test_zip.writestr(xlsx_path, string_io.read())
# test_zip.writestr(xlsx_path, workbook.data_only)
test_zip.close()


