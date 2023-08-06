"""
Allows a user to open a workbook, select cells they don't want to share the data for,
and opens the new spreadsheet with only non-specific data about the selected cell values

Example: suppose the user has some names which need to be kept anonymous. These
names are selected, and the workbook we receive only contains information such 
as the length of the names, and whether the names are alphanumeric strings


TO-DO: make this work for Excel sheets stored non-locally.

TO-DO. Obviously we need to ask user for where they want the save location,
rather than hardcoding it in here! and maybe ask if they want us to open it afterwards


Also TO-DO: this requires the version of the workbook the user wants to perform this
on to be saved, so maybe a notification for this
"""

save_location = r'C:\Users\ethan\Documents\heatmap.xlsx'


import xlwings as xw
import openpyxl 
import io
wb_xw= xw.apps.active.books.active

from index_helpers import convert_from_tuple, block_to_list, convert_to_tuple


user_selection_address = wb_xw.selection.address
user_selection_address_by_block = user_selection_address.split(",")
cells_to_hide = set()
for block in user_selection_address_by_block:
    cells = block_to_list(block).split(",")
    for cell in cells:
        cells_to_hide.add(convert_to_tuple(cell))


#note this currently only works for excel files stored locally
xlsx_filename=wb_xw.fullname
with open(xlsx_filename, "rb") as f:
    #this needs to be done as well as wb.close to make sure the workbook doesnt get broken by saving and other actions
    
    in_mem_file = io.BytesIO(f.read())

wb_pyxl = openpyxl.load_workbook(in_mem_file, read_only=True)

sheet_name = wb_xw.sheets.active.name

ws_pyxl = wb_pyxl[sheet_name]

def generic_string_info(input_string):
    return {"Whitespaces": " ".count(input_string),
           "Length": len(input_string),
           "alpha only?": input_string.isalpha(),
           "numeric only?": input_string.isnumeric(),
           "alpha num only?": input_string.isalnum(),
           "non alpha num": [char for char in input_string if not char.isalnum()]}


wb = openpyxl.Workbook(write_only=True)
ws = wb.create_sheet()

ws_row = ws_pyxl.max_row
ws_col = ws_pyxl.max_column
fill_for_hidden_cells = openpyxl.styles.PatternFill(fill_type="solid",
                 start_color='FFFFFF00',
                 end_color='FF000000')

for j in range(1, ws_row+1):
    #j counts along the rows, 1,2,3,...
    row = []

    for i in range(1, ws_col+1):
        #i counts along the columns, i.e. A,B,C,D,...
        cell=openpyxl.cell.WriteOnlyCell(ws)
        if (i,j) in cells_to_hide:
            if ws_pyxl[convert_from_tuple((i,j))].value != None:
                cell.fill = fill_for_hidden_cells
                value_in_sheet = ws_pyxl[convert_from_tuple((i,j))].value
                cell.value = str(type(value_in_sheet))
                if type(value_in_sheet) == str:
                    cell_comment = openpyxl.comments.Comment(str(generic_string_info(value_in_sheet)), "Summary data bot")
                    cell.comment = cell_comment
                
            
            
        else:
            cell.value = ws_pyxl[convert_from_tuple((i,j))].value 
        row.append(cell)
    print(row)
    ws.append(row)
wb.save(save_location)       


xw.Book(save_location) #open the output