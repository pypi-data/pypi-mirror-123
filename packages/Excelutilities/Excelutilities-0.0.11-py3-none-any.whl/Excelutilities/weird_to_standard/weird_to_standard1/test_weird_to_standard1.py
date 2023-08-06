import unittest
import openpyxl
import os
import io
from Excelutilities.weird_to_standard.weird_to_standard1.weird_to_standard1 import weird_to_standard1_array
from Excelutilities.cleaning_utilities.worksheet_cleaning_utilities import remove_empty_rows_and_columns_input_ws_output_val_list

class TestStringMethods(unittest.TestCase):

    def test_weird_to_standard1_array(self, load_workbook = openpyxl.load_workbook, BytesIO= io.BytesIO):
        wb_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),"wts_1_preprocessing_example.xlsx")
        xlsx_filename=wb_path
        with open(xlsx_filename, "rb") as f:
            #this needs to be done as well as wb.close to make sure the workbook doesnt get broken by saving and other actions
            
            in_mem_file = BytesIO(f.read())

        wb = load_workbook(in_mem_file, read_only=True)
        ws = wb["Sheet1"]
        color_hex = "FFFF0000"
        ws_block= ws["A2:D15"]
        return_array = weird_to_standard1_array(ws_block, color_hex, search_breadth=2)
        wb.close()

        wb_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),"wts_2_postprocessing_example.xlsx")
        xlsx_filename=wb_path
        with open(xlsx_filename, "rb") as f:
            #this needs to be done as well as wb.close to make sure the workbook doesnt get broken by saving and other actions
            
            in_mem_file = BytesIO(f.read())
        wb = load_workbook(in_mem_file, read_only=True)
        ws = wb["Sheet1"]
        test_array = remove_empty_rows_and_columns_input_ws_output_val_list(ws)

        wb.close()
        
        self.assertEqual(return_array,test_array)
if __name__ == '__main__':
    unittest.main()