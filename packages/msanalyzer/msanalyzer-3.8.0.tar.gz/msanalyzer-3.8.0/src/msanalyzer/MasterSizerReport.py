from __future__ import annotations

import io
import os
import warnings
from datetime import date
from enum import Enum, unique

warnings.filterwarnings("ignore", "(?s).*MATPLOTLIBDATA.*", category=UserWarning)
import logging
from collections.abc import Callable
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import numpy.typing as npt
import openpyxl

from . import __version__
from .MasterSizerInput import MasterSizerInput
from .models.SizeDistributionBaseModel import SizeDistributionBaseModel
from .models.SizeDistributionModelsFactory import getPSDModelsList

logger = logging.getLogger(__name__)
__author__: str = "Marcus Bruno Fernandes Silva"
__email__: str = "marcusbfs@gmail.com"


@dataclass
class ParametersData:
    repr: str
    value: float
    stddev: float


@dataclass
class ModelData:
    name: str
    expr: str
    parameters: List[ParametersData]
    r2: float
    s: float
    D: Dict[str, float]


@dataclass
class BestData:
    s: str
    r2: str


@dataclass
class ModelsData:
    models: List[ModelData]
    best: BestData


@unique
class DiameterMeanType(Enum):
    geometric = 1
    arithmetic = 2


class MasterSizerReport:
    def __init__(self, input_reader: MasterSizerInput = MasterSizerInput()) -> None:
        self.__diameters_filename: str = ""
        self.__vol_in_per_filename: str = ""
        self.__number_of_points: int = 0
        self.__x_data: np.ndarray = np.array([])
        self.__x_data_geomean: np.ndarray = np.array([])
        self.__x_data_aritmeticmean: np.ndarray = np.array([])
        self.__y_data: np.ndarray = np.array([])
        self.__x_data_mean: np.ndarray = np.array([])
        self.__cumulative_y_vals: np.ndarray = np.array([])
        self.__diff_of_cumulative_y_vals: np.ndarray = np.array([])
        self.__ms_input: MasterSizerInput = input_reader
        self.__version: str = __version__
        self.__input_xps_file: str = ""
        self.__meantype: DiameterMeanType = DiameterMeanType.geometric
        self.__headers: List[str] = [
            "diameter [microns]",
            "volume fraction [-]",
            "cumulative volume fraction [-]",
        ]
        self.__log_scale: bool = False
        self.__models: List[SizeDistributionBaseModel] = getPSDModelsList()
        self.__num_of_models: int = len(self.__models)

    # Public
    def setLogScale(self, logscale: bool = True) -> None:
        self.__log_scale = logscale
        logger.info("Log scale setted to {}".format(logscale))

    def setXPSfile(self, xps_mem: io.BytesIO, xps_filename: str) -> None:
        self.__input_xps_file = xps_filename
        logger.info('XPS file setted to "{}"'.format(xps_filename))
        self.__ms_input.setFile(xps_mem, xps_filename)
        self.__updateXY_data()

    def setXandY(self, x_vals: npt.ArrayLike, y_vals: npt.ArrayLike) -> None:
        self.__x_data = np.array(x_vals)
        self.__y_data = np.array(y_vals)
        self.__number_of_points = len(self.__y_data)
        assert len(self.__x_data) == len(self.__y_data) + 1

    def genCumulativeSizeDistribution(self) -> None:
        self.__cumulative_y_vals = np.zeros(self.__number_of_points)
        for i in range(1, self.__number_of_points):
            self.__cumulative_y_vals[i] = (
                self.__y_data[i] + self.__cumulative_y_vals[i - 1]
            )

    def genDiffCumulativeSizeDistribution(self) -> None:
        self.__diff_of_cumulative_y_vals = np.diff(
            self.__cumulative_y_vals, prepend=0.0
        ) / np.diff(self.__x_data_mean, prepend=1.0)

    # saving files

    def saveExcel(self, output_dir: str, base_filename: str) -> None:
        filename = os.path.join(output_dir, base_filename + ".xlsx")

        wb = openpyxl.Workbook()
        ws1: openpyxl.Workbook = wb.active

        for i, header in enumerate(self.__headers):
            column_width = len(header) + 1
            ws1.cell(row=1, column=i + 1, value=header).font = openpyxl.styles.Font(
                bold=True
            )

            ws1.cell(row=1, column=i + 1).alignment = openpyxl.styles.Alignment(
                horizontal="right"
            )

            ws1.column_dimensions[
                openpyxl.utils.get_column_letter(i + 1)
            ].width = column_width

        for i in range(len(self.__x_data_mean)):
            ws1.cell(row=i + 2, column=1, value=self.__x_data_mean[i])
            ws1.cell(row=i + 2, column=2, value=self.__y_data[i])
            ws1.cell(row=i + 2, column=3, value=self.__cumulative_y_vals[i])

        last_row: int = len(self.__x_data_mean) + 1

        # Charts

        xvalues = openpyxl.chart.Reference(ws1, min_col=1, min_row=2, max_row=last_row)

        def _genChart(
            y_title: str,
            y_col: int,
            title: str,
            y_scale_bound_to_one: bool = False,
        ) -> openpyxl.chart.ScatterChart:
            chart = openpyxl.chart.ScatterChart()
            chart.style = 13
            chart.title = y_title
            chart.x_axis.title = "Diameter [microns]"
            chart.y_axis.title = ""

            chart.legend = None

            if y_scale_bound_to_one:
                # chart.x_axis.scaling.min = 0
                # chart.x_axis.scaling.max = 11

                chart.y_axis.scaling.min = 0
                chart.y_axis.scaling.max = 1.1

            values = openpyxl.chart.Reference(
                ws1, min_col=y_col, min_row=2, max_row=last_row
            )
            series = openpyxl.chart.Series(values, xvalues, title_from_data=True)
            chart.append(series)
            return chart

        chart1 = _genChart("Volume fraction [-]", 2, "")
        ws1.add_chart(chart1, "E2")

        chart2 = _genChart(
            "Cumulative volume fraction[-]", 3, "", y_scale_bound_to_one=True
        )
        ws1.add_chart(chart2, "E17")

        wb.save(filename=filename)
        logger.info('Exported data to excel file: "{}"'.format(filename))
        return

    def saveFig(self, output_dir: str, base_filename: str) -> plt.figure:
        # TODO: fix this. we are calling this function way too much!
        # self.genDiffCumulativeSizeDistribution()

        # print(self.getDiffOfCumulativeYvalues())

        # clear old figures
        plt.clf()
        # plot
        fig, ax1 = plt.subplots()
        ax2 = plt.twinx()
        ax1.set_ylabel("volume fraction (dX) [-]")
        ax2.set_ylabel("cumulative distribution (X) [-]")
        ax2.grid()

        ax1_color = "#1f77b4"
        ax2_color = "red"

        ax1.tick_params(axis="y", colors=ax1_color, which="both")
        ax1.plot(
            self.getXmeanValues(),
            self.getYvalues(),
            linestyle="--",
            marker="o",
            color=ax1_color,
            label="dX",
        )

        ax2.tick_params(axis="y", colors=ax2_color, which="both")
        ax2.plot(
            self.getXmeanValues(),
            self.getCumulativeYvalues(),
            linestyle="--",
            marker="o",
            color=ax2_color,
            label="X",
        )

        if self.__log_scale:
            ax1.set_xlabel("log scale - diameter [$\mu m$]")
            self.__format_LogScale_Xaxis(ax1)
        else:
            ax1.set_xlabel("diameter [$\mu m$]")

        filename = os.path.join(output_dir, base_filename + ".svg")
        plt.savefig(filename, dpi=1200)
        logger.info('Saved curves to "{}"'.format(filename))
        return fig
        # end of plot

    def saveModelsFig(
        self, output_dir: str, base_filename: str, callback: Callable[[], None] = None
    ) -> dict:

        models_figs = {}

        for model in self.__models:
            # plot
            fig, ax = plt.subplots()

            ax.set_ylabel("Cumulative distribution (X) [-]")
            ax.grid()

            ax.plot(
                self.__x_data_mean,
                model.compute(self.__x_data_mean),
                label="{} Model".format(model.getModelName()),
                linestyle="--",
                color="red",
            )
            ax.scatter(
                self.__x_data_mean,
                self.__cumulative_y_vals,
                label="Data",
                facecolors="none",
                edgecolors="black",
            )

            ax.legend()

            if self.__log_scale:
                ax.set_xlabel("log scale - diameter [$\mu m$]")
                self.__format_LogScale_Xaxis(ax)
            else:
                ax.set_xlabel("diameter [$\mu m$]")

            filename = os.path.join(
                output_dir, model.getModelName() + "_" + base_filename + ".svg"
            )
            plt.savefig(filename, dpi=1200)
            logger.info('Saved {} curve to "{}"'.format(model.getModelName(), filename))

            models_figs[model.getModelName()] = fig
            # end of plot

            # call callback - maybe to update a progress bar
            if callback is not None:
                callback()

        return models_figs

    def saveData(self, output_dir: str, data_filename: str) -> None:
        output_file = os.path.join(output_dir, data_filename)
        header = "%10s\t%10s\t%10s" % (
            self.__headers[0],
            self.__headers[1],
            self.__headers[2],
        )
        np.savetxt(
            output_file,
            np.transpose(
                np.array([self.__x_data_mean, self.__y_data, self.__cumulative_y_vals])
            ),
            fmt="%15.10f",
            header=header,
        )

        content = self.getTxtFilesHeader()
        with open(output_file, "r+") as f:
            content += f.read()

        with open(output_file, "w") as f:
            f.write(content)

        logger.info('Saved curves data to "{}"'.format(output_file))

    def saveModelsData(self, output_dir: str, data_filename: str) -> None:
        for model in self.__models:
            output_file = os.path.join(
                output_dir, model.getModelName() + "_" + data_filename + ".txt"
            )
            content = self.getTxtFilesHeader()
            content += model.getFormattedOutput()
            with open(output_file, "w") as of:
                of.write(content)
            logger.info(
                'Saved {} data to "{}"'.format(model.getModelName(), output_file)
            )
        return

    def saveBestModelsRanking(self, output_dir: str, base_filename: str) -> None:
        output_file = os.path.join(output_dir, base_filename + ".txt")

        file_content = self.getTxtFilesHeader()
        file_content += "Best models\n"
        file_content += "===========\n\n"

        file_content += (
            "Ranking based on S (standard error mean). Lowest value is better.\n\n"
        )

        i = 1
        for model in self.ranking_models_S_based:
            file_content += "  #{:02d} - {}: S = {:.7f}\n".format(
                i, model.getModelName(), model.getStdErrorMean()
            )
            i += 1

        file_content += "\n"
        file_content += (
            "Ranking based on R-squared (not trustworthy). Highest value is better.\n\n"
        )

        i = 1
        for model in self.ranking_models_r2_based:
            file_content += "  #{:02d} - {}: R-squared = {:.7f}\n".format(
                i, model.getModelName(), model.getRsquared()
            )
            i += 1

        with open(output_file, "w") as of:
            of.write(file_content)
        logger.info('Saved best model rankings to "{}"'.format(output_file))

    def evaluateData(self) -> None:
        # Extract data
        self.__number_of_points = len(self.__y_data)

        # Get x values means
        self.__x_data_geomean = np.sqrt(self.__x_data[1:] * self.__x_data[:-1])
        self.__x_data_aritmeticmean = 0.5 * (self.__x_data[1:] + self.__x_data[:-1])

        # Use geometric mean for default values of x
        self.setDiameterMeanType(self.__meantype)

        self.genCumulativeSizeDistribution()
        self.genDiffCumulativeSizeDistribution()

    def getNumberOfModels(self) -> int:
        return len(self.__models)

    def evaluateModels(self, callback: Callable[[], None] = None) -> ModelsData:

        self.genDiffCumulativeSizeDistribution()

        for model in self.__models:
            model.evaluate(self.__x_data_mean, self.__cumulative_y_vals)
            if callback is not None:
                callback()

        logger.info(
            "Searching best R-squared (highest value) and best S (lowest value) models"
        )

        self.ranking_models_S_based: List[SizeDistributionBaseModel] = sorted(
            self.__models, key=lambda x: x.getStdErrorMean()
        )
        logger.info("Ranking based on S: {}".format(self.ranking_models_S_based))
        self.ranking_models_r2_based: List[SizeDistributionBaseModel] = sorted(
            self.__models, key=lambda x: x.getRsquared(), reverse=True
        )
        logger.info(
            "Ranking based on R-squared: {}".format(self.ranking_models_r2_based)
        )

        ret: Dict[str, Any] = {"models": []}

        models = ModelsData(models=[], best=BestData(s="", r2=""))

        for model in self.ranking_models_S_based:
            par_str = model.getParametersStr()
            par_val = model.getParametersValues()
            par_stddev = model.getParametersStdDev()

            m = ModelData(
                name=model.getModelName(),
                expr=model.getExpression(),
                parameters=[],
                r2=model.getRsquared(),
                s=model.getStdErrorMean(),
                D={},
            )

            for i in range(len(par_str)):
                par = ParametersData(
                    repr=par_str[i], value=par_val[i], stddev=par_stddev[i]
                )
                m.parameters.append(par)

            for n in (10, 25, 50, 75, 90):
                m.D["D" + str(n)] = model.getDnFromCompute(n / 100.0)

            models.models.append(m)

        best = BestData(
            s=self.ranking_models_S_based[0].getModelName(),
            r2=self.ranking_models_r2_based[0].getModelName(),
        )
        models.best = best

        return models

    def cutLastNPoints(self, number_of_points: int) -> None:
        if number_of_points > 0:
            self.__x_data = self.__x_data[:-number_of_points].copy()
            self.__y_data = self.__y_data[:-number_of_points].copy()
            self.__number_of_points = len(self.__y_data)
            logger.info("Cutted the last {} null points".format(number_of_points))
            logger.info("New length of x: {}".format(len(self.__x_data)))
        else:
            logger.info("No points ned to be cutted")
            logger.info("length of x: {}".format(len(self.__x_data)))

    def cutFirstNPoints(self, number_of_points: int) -> None:
        if number_of_points > 0:
            self.__x_data = self.__x_data[number_of_points:].copy()
            self.__y_data = self.__y_data[number_of_points:].copy()
            self.__number_of_points = len(self.__y_data)
            logger.info("Cutted the first {} null points".format(number_of_points))
            logger.info("New length of x: {}".format(len(self.__x_data)))
        else:
            logger.info("No points ned to be cutted")
            logger.info("length of x: {}".format(len(self.__x_data)))

    def cutLastZeroPoints(
        self, number_of_lefting_zeros: int, tol: float = 1e-10
    ) -> None:
        total_zeros = 0

        for i in range(self.__number_of_points - 1, -1, -1):
            if self.__isFloatEqual(0.0, self.__y_data[i], tol=tol):
                total_zeros += 1
            else:
                break

        if total_zeros >= number_of_lefting_zeros:
            self.cutLastNPoints(total_zeros - number_of_lefting_zeros - 1)

    def cutFirstZeroPoints(
        self, number_of_lefting_zeros: int, tol: float = 1e-10
    ) -> None:
        total_zeros = 0

        for i in range(self.__number_of_points):
            if self.__isFloatEqual(0.0, self.__y_data[i], tol=tol):
                total_zeros += 1
            else:
                break

        if total_zeros >= number_of_lefting_zeros:
            self.cutFirstNPoints(total_zeros - number_of_lefting_zeros - 1)

    # Getters
    def getNumberOfPoints(self) -> int:
        return self.__number_of_points

    def getRRBParameters(self) -> Tuple[float, float]:
        for model in self.__models:
            model_name = model.getModelName()
            if model_name.lower() == "rrb":
                pars = model.getParametersValues()
                print(pars)
                return (pars[0], pars[1])

        raise RuntimeError("RRB parameters not found")

    def getGeometricMeanXvalues(self) -> np.ndarray:
        return self.__x_data_geomean

    def getAritmeticMeanXvalues(self) -> np.ndarray:
        return self.__x_data_geomean

    def getRawXvalues(self) -> np.ndarray:
        return self.__x_data

    def getCumulativeYvalues(self) -> np.ndarray:
        return self.__cumulative_y_vals

    def getDiffOfCumulativeYvalues(self) -> np.ndarray:
        return self.__diff_of_cumulative_y_vals

    def getYvalues(self) -> np.ndarray:
        return self.__y_data

    def getXmeanValues(self) -> np.ndarray:
        return self.__x_data_mean

    def getTxtFilesHeader(self) -> str:
        content = ""
        content += " msanalyzer {} \n\n".format(__version__)
        content += " Author: {} \n".format(__author__)
        content += " email: {} \n\n".format(__email__)
        content += ' file analyzed: "{}" \n'.format(
            os.path.abspath(self.__input_xps_file)
        )
        content += " Date: {} \n".format(date.today().strftime("%d-%b-%Y"))
        content = self.getBorderedText(content)
        content += "\n\n"
        return content

    def getBorderedText(self, text: str) -> str:
        lines = text.splitlines()
        width = max(len(s) for s in lines)
        res = ["+" + "-" * width + "+"]
        for s in lines:
            res.append("|" + (s + " " * width)[:width] + "|")
        res.append("+" + "-" * width + "+")
        return "\n".join(res)

    def getInputFile(self) -> str:
        return self.__input_xps_file

    @staticmethod
    def getVersion() -> str:
        return __version__

    # Setters
    def setGeometricMean(self) -> None:
        self.__x_data_mean = self.__x_data_geomean

    def setArithmeticMean(self) -> None:
        self.__x_data_mean = self.__x_data_aritmeticmean

    def setDiameterMeanType(self, typed: DiameterMeanType) -> None:
        self.__meantype = typed
        if DiameterMeanType.geometric == typed:
            self.setGeometricMean()
        elif DiameterMeanType.arithmetic == typed:
            self.setArithmeticMean()
        logger.info("Diameter mean type setted to {}".format(typed))

    # Private

    @staticmethod
    def formatLogScaleXaxis(xaxis: matplotlib.axes.Axes) -> None:
        from matplotlib.ticker import ScalarFormatter

        xaxis.set_xscale("log")
        xaxis.set_xlabel("log scale - diameter [$\mu m$]")
        for axis in [xaxis.xaxis, xaxis.yaxis]:
            axis.set_major_formatter(ScalarFormatter())
        return

    def __format_LogScale_Xaxis(self, xaxis: matplotlib.axes.Axes) -> None:
        return MasterSizerReport.formatLogScaleXaxis(xaxis)

    def __updateXY_data(self) -> None:
        self.__ms_input.extractData()
        self.__x_data = self.__ms_input.getx()
        self.__y_data = self.__ms_input.gety()
        self.__number_of_points = len(self.__y_data)

    def __setVolInPerFile(self, filename: str) -> None:
        self.__vol_in_per_filename = filename

    def __setDiametersFile(self, filename: str) -> None:
        self.__diameters_filename = filename

    def __isFloatEqual(self, x: float, y: float, tol: float = 1e-10) -> bool:
        return bool(np.abs(x - y) < tol)
